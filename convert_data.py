#!/usr/bin/env python3
"""Convert all batch Excel files to unified JSON for the SAM web app."""

import openpyxl
import json
import os
import re
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = '/home/shahana-v/Downloads/drive-download-20260224T095615Z-1-001'
OUTPUT_DIR = os.path.join(BASE_DIR, 'data')

DEPT_NAMES = {
    'ATE': 'Automobile Engineering',
    'CSE': 'Computer Science and Engineering',
    'CVE': 'Civil Engineering',
    'DSC': 'Data Science',
    'ECE': 'Electronics and Communication Engineering',
    'EEE': 'Electrical and Electronics Engineering',
    'IMT': 'Information Technology',
    'MCE': 'Mechanical Engineering',
}

def extract_students(filepath, dept_code, batch_year):
    """Extract student records from an Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Student_Roster']
    
    students = []
    for row in range(2, ws.max_row + 1):
        student_id = ws.cell(row, 1).value
        if not student_id or not str(student_id).strip():
            continue
        
        name = ws.cell(row, 4).value
        gender = ws.cell(row, 5).value
        email = ws.cell(row, 6).value
        cgpa = ws.cell(row, 7).value
        
        if not name or not str(name).strip():
            continue
        
        gender_raw = str(gender).strip().upper() if gender else 'U'
        if gender_raw.startswith('M'):
            gender_norm = 'M'
        elif gender_raw.startswith('F'):
            gender_norm = 'F'
        else:
            gender_norm = 'U'
        
        cgpa_val = None
        try:
            if cgpa and str(cgpa).strip() and str(cgpa).strip().lower() not in ('nil', 'na', 'n/a', '-'):
                cgpa_val = float(cgpa)
        except (ValueError, TypeError):
            pass
        
        students.append({
            'id': str(student_id).strip(),
            'name': str(name).strip(),
            'gender': gender_norm,
            'email': str(email).strip() if email else '',
            'cgpa': cgpa_val,
            'department': dept_code,
            'batch': batch_year,
        })
    
    wb.close()
    return students

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    all_data = {
        'college': 'Government College of Engineering, Erode',
        'batches': [],
    }
    
    # Auto-detect batch folders (directories named with 4 digits)
    batch_dirs = sorted([
        d for d in os.listdir(DATA_DIR)
        if os.path.isdir(os.path.join(DATA_DIR, d)) and re.match(r'^\d{4}$', d)
    ])
    
    print(f"Found batch folders: {batch_dirs}")
    
    grand_total = 0
    
    for batch_dir in batch_dirs:
        batch_year = int(batch_dir)
        batch_path = os.path.join(DATA_DIR, batch_dir)
        
        # Auto-detect department Excel files
        xlsx_files = sorted([
            f for f in os.listdir(batch_path)
            if f.endswith('.xlsx') and f.startswith(f'{batch_year}_')
        ])
        
        batch_data = {
            'year': batch_year,
            'departments': [],
            'students': [],
        }
        
        print(f"\n--- Batch {batch_year} ({len(xlsx_files)} files) ---")
        
        for xlsx_file in xlsx_files:
            # Extract dept code: e.g. "2027_CSE.xlsx" -> "CSE"
            dept_code = xlsx_file.replace(f'{batch_year}_', '').replace('.xlsx', '')
            dept_name = DEPT_NAMES.get(dept_code, dept_code)
            
            filepath = os.path.join(batch_path, xlsx_file)
            students = extract_students(filepath, dept_code, batch_year)
            males = sum(1 for s in students if s['gender'] == 'M')
            females = sum(1 for s in students if s['gender'] == 'F')
            
            batch_data['departments'].append({
                'code': dept_code,
                'name': dept_name,
                'totalStudents': len(students),
                'males': males,
                'females': females,
            })
            
            batch_data['students'].extend(students)
            print(f"  {dept_code}: {len(students)} students (M:{males}, F:{females})")
        
        batch_total = len(batch_data['students'])
        grand_total += batch_total
        print(f"  Batch {batch_year} total: {batch_total} students")
        
        all_data['batches'].append(batch_data)
    
    print(f"\n=== Grand Total: {grand_total} students across {len(batch_dirs)} batches ===")
    
    # Save JSON
    json_path = os.path.join(OUTPUT_DIR, 'students.json')
    with open(json_path, 'w') as f:
        json.dump(all_data, f, indent=2)
    print(f"Saved JSON to {json_path}")
    
    # Save JS (for file:// CORS workaround)
    js_path = os.path.join(OUTPUT_DIR, 'students_data.js')
    with open(js_path, 'w') as f:
        f.write('// Auto-generated — all batch data for SAM webapp\n')
        f.write('const STUDENT_DATA = ')
        json.dump(all_data, f)
        f.write(';\n')
    print(f"Saved JS to {js_path}")

if __name__ == '__main__':
    main()
